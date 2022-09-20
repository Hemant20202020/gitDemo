import openpyxl
import pytest

#import csv
from HT_Function import HT
from openpyxl.styles import PatternFill
from openpyxl import Workbook
wb_1=Workbook()
@pytest.fixture(scope='session')
def final_valid_value():
    sh1=wb_1.active
    lst=['AL','LP','LQ','LR','LS','NO']
    sh1.append(lst)
    for i in range(1, 2):
        for j in range(1, 7):
            sh1.cell(row=i, column=j).fill = PatternFill('solid', fgColor='ff5733')
    yield sh1

    wb_1.save('final_valid_value.xlsx')

wb=openpyxl.load_workbook('valid_value_1.xlsx')
sh=wb['Sheet']
rows=sh.max_row
cols=sh.max_column
lst=[]

for row in range(2,rows+1):
    a=[]
    for col in range(1,cols+1):
        a.append(sh.cell(row,col).value)

    lst.append(a)

@pytest.mark.parametrize('AL,setup_LP,setup_LQ,setup_LR,setup_LS,NO',[tuple(x) for x in lst])
def test_data_set(AL,setup_LP,setup_LQ,setup_LR,setup_LS,NO,final_valid_value):
    if (setup_LP <=0 or setup_LP >=8) or (setup_LQ <=7 or setup_LQ >= 16) or (setup_LR <=1 or setup_LR >=10) or (
            setup_LS <=9 or setup_LS >=18):

        pytest.fail('Invalid Range')
    else:
        lst_01=[AL,setup_LP,setup_LQ,setup_LR,setup_LS,NO]
        final_valid_value.append(lst_01)
        assert HT(AL, setup_LP, setup_LQ, setup_LR, setup_LS, NO)

