import openpyxl
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
"""c1 = my_sheet.cell(row = 1, column = 1)
c1.value = "Aadrika"
c2 = my_sheet.cell(row= 1 , column = 2)
c2.value = "Adwaita"
c3 = my_sheet['A2']
c3.value = "Satyajit"
# B2 = column = 2 & row = 2.
c4 = my_sheet['B2']
c4.value = "Bivas" """
for i in range(1,10):
    c1=my_sheet.cell(row=i,column=1)
    c1.value=i
my_wb.save("Book1.xlsx")