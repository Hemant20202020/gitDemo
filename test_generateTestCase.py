import time
import openpyxl
import pytest
from openpyxl import Workbook
from HT_Function import HT
from openpyxl.styles import PatternFill
from Token_Bucket_Algo import TokenBucket
wb=Workbook()
@pytest.fixture(scope='session')
def valid_value():
    sh1=wb.active
    lst=['AL','LP','LQ','LR','LS','NO']
    sh1.append(lst)
    for i in range(1, 2):
        for j in range(1, 7):
            sh1.cell(row=i, column=j).fill = PatternFill('solid', fgColor='ff5733')
    yield sh1
    #for i in range(2,386):
     #   sh1.cell(row=i,column=7).fill=PatternFill('solid',fgColor='7aff33')
    wb.save('valid_value_1.xlsx')

LP=[i for i in range(1,8)]
LP.extend([-10,21])
LQ=[i for i in range(8,16)]
LQ.extend([7,20])
LR=[i for i in range(2,10)]
LR.extend([1,10])
LS=[i for i in range(10,18)]
LS.extend([9,15])


@pytest.fixture(params=LP)
def setup_LP(request):
    return request.param

@pytest.fixture(params=LQ)
def setup_LQ(request):
    return request.param

@pytest.fixture(params=LR)
def setup_LR(request):
    return request.param

@pytest.fixture(params=LS)
def setup_LS(request):
    return request.param

@pytest.fixture(scope='session')
def setup_token_bucket():
    def forward(packet):
        print("TestCase Forwarded: " + str(packet))
    def drop(packet):
        print("TestCase Dropped: " + str(packet))
    tbucket=TokenBucket(1,1,forward,drop)
    yield tbucket

@pytest.mark.parametrize('AL,NO',[(1.0,2)])
def test_create_dataset(AL,setup_LP,setup_LQ,setup_LR,setup_LS,NO,valid_value,setup_token_bucket):
        lst = [AL, setup_LP, setup_LQ, setup_LR, setup_LS, NO]
        packet=lst
        time.sleep(0.2)
        forward_packet=setup_token_bucket.handle(packet)
        if forward_packet[0]!=None:
            AL_1 = forward_packet[0]
            setup_LP_1 = forward_packet[1]
            setup_LQ_1 = forward_packet[2]
            setup_LR_1 = forward_packet[3]
            setup_LS_1 = forward_packet[4]
            NO_1 = forward_packet[5]
            lst_1 = [AL_1, setup_LP_1, setup_LQ_1, setup_LR_1, setup_LS_1, NO_1]
            valid_value.append(lst_1)


